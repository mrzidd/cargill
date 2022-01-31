from datetime import datetime
import pandas as pd
import xlrd
from openpyxl import load_workbook


class Stock_Barco:
    def __init__(self, StockID, Make, Model,ColorID,VehicleType ,CostPrice,SpareParts,LaborCost ,Registration_Date ,Mileage ,PurchaseDate,VehicleAgeInYears):
        self.StockID = StockID
        self.Make = Make
        self.Model = Model
        self.ColorID = ColorID
        self.VehicleType = VehicleType
        self.CostPrice = CostPrice
        self.SpareParts = SpareParts
        self.LaborCost = LaborCost
        self.Registration_Date = Registration_Date
        self.Mileage = Mileage
        self.PurchaseDate = PurchaseDate
        self.VehicleAgeInYears = VehicleAgeInYears

    def __eq__ (self, other):
        return self.StockID == other.StockID

    def __str__(self):
        return "StockID: " + str(self.StockID) + "\nMake: " + str(self.Make) + "\nModel: " + str(self.Model) + "\nColorID: " + str(self.ColorID) + "\nVehicleType: " + str(self.VehicleType) + "\nCostPrice: " + str(self.CostPrice) + "\nSpareParts: " + str(self.SpareParts) + "\nLaborCost: " + str(self.LaborCost) + "\nRegistration_Date: " + str(self.Registration_Date) + "\nMileage: " + str(self.Mileage) + "\nPurchaseDate: " + str(self.PurchaseDate) + "\nVehicleAgeInYears: " + str(self.VehicleAgeInYears)

    def get_StockID(self):
        return self.StockID
    
    def get_Make(self):
        return self.Make

    def get_Model(self):
        return self.Model

    def get_ColorID(self):
        return self.ColorID

    def get_VehicleType(self):
        return self.VehicleType

    def get_CostPrice(self):
        return self.CostPrice

    def get_SpareParts(self):
        return self.SpareParts

    def get_LaborCost(self):
        return self.LaborCost

    def get_Registration_Date(self):
        return self.Registration_Date

    def get_Mileage(self):
        return self.Mileage

    def get_PurchaseDate(self):
        return self.PurchaseDate

    def get_VehicleAgeInYears(self):
        return self.VehicleAgeInYears

    def set_StockID(self, StockID):
        self.StockID = StockID

    def set_Make(self, Make):
        self.Make = Make

    def set_Model(self, Model):
        self.Model = Model

    def set_ColorID(self, ColorID):
        self.ColorID = ColorID

    def set_VehicleType(self, VehicleType):
        self.VehicleType = VehicleType

    def set_CostPrice(self, CostPrice):
        self.CostPrice = CostPrice

    def set_SpareParts(self, SpareParts):
        self.SpareParts = SpareParts

    def set_LaborCost(self, LaborCost):
        self.LaborCost = LaborCost

    def set_Registration_Date(self, Registration_Date):
        self.Registration_Date = Registration_Date

    def set_Mileage(self, Mileage):
        self.Mileage = Mileage

    def set_PurchaseDate(self, PurchaseDate):
        self.PurchaseDate = PurchaseDate

    def set_VehicleAgeInYears(self, VehicleAgeInYears):
        self.VehicleAgeInYears = VehicleAgeInYears

    
    def create_stock_list(self):
   

        # ask user for StockID, Make, Model,ColorID,VehicleType ,CostPrice,SpareParts,LaborCost ,Registration_Date ,Mileage ,PurchaseDate,VehicleAgeInYears
        StockID = self.get_StockID()
        Make = self.get_Make()
        Model = self.get_Model()
        ColorID = self.get_ColorID()
        VehicleType = self.get_VehicleType()
        CostPrice = self.get_CostPrice()
        SpareParts = self.get_SpareParts()
        LaborCost = self.get_LaborCost()
        Registration_Date = self.get_Registration_Date()
        Mileage = self.get_Mileage()
        PurchaseDate = self.get_PurchaseDate()
        VehicleAgeInYears = self.get_VehicleAgeInYears()

        diction = {'StockID': StockID, 'Make': Make, 'Model': Model, 'ColorID': ColorID, 'VehicleType': VehicleType, 'CostPrice': CostPrice, 'SpareParts': SpareParts, 'LaborCost': LaborCost, 'Registration_Date': Registration_Date, 'Mileage': Mileage, 'PurchaseDate': PurchaseDate, 'VehicleAgeInYears': VehicleAgeInYears}

        lista = [[StockID, Make, Model, ColorID, VehicleType, CostPrice, SpareParts, LaborCost, Registration_Date, Mileage, PurchaseDate, VehicleAgeInYears]]
        # append the dictionary to dataframe
        df = pd.DataFrame(diction, index=[0])
        print(df)

        book = load_workbook('cargill.xlsx')

        if 'barco' in book.sheetnames:
            print('sheet1 exists')
            page = book['barco']
        else:
            # create a new sheet
            page = book.create_sheet('barco')
            page = book['barco']

        for row in lista:
            page.append(row)
        book.save('cargill.xlsx')



# create class datedimension with attributes DateKey,Year,MonthNum,MonthFull,MonthAbbr,QuarterNum,QuarterFull,QuarterAbbr,YearAndQuarterNum,QuarterAbbrAndYear,MonthAbbrAndYear,MonthAndYear,MonthName,MonthNameAbbr,QuarterAndYear,QuarterAndYearAbbr2,YearAndMonthNum
class DateDimension:
    def __init__(self, DateKey, Year, MonthNum, MonthFull, MonthAbbr, QuarterNum, QuarterFull, QuarterAbbr, YearAndQuarterNum, QuarterAbbrAndYear, MonthAbbrAndYear, MonthAndYear, MonthName, MonthNameAbbr, QuarterAndYear, QuarterAndYearAbbr2, YearAndMonthNum):
        self.DateKey = DateKey
        self.Year = Year
        self.MonthNum = MonthNum
        self.MonthFull = MonthFull
        self.MonthAbbr = MonthAbbr
        self.QuarterNum = QuarterNum
        self.QuarterFull = QuarterFull
        self.QuarterAbbr = QuarterAbbr
        self.YearAndQuarterNum = YearAndQuarterNum
        self.QuarterAbbrAndYear = QuarterAbbrAndYear
        self.MonthAbbrAndYear = MonthAbbrAndYear
        self.MonthAndYear = MonthAndYear
        self.MonthName = MonthName
        self.MonthNameAbbr = MonthNameAbbr
        self.QuarterAndYear = QuarterAndYear
        self.QuarterAndYearAbbr2 = QuarterAndYearAbbr2
        self.YearAndMonthNum = YearAndMonthNum

    def get_DateKey(self):
        return self.DateKey

    def get_Year(self):
        return self.Year

    def get_MonthNum(self):
        return self.MonthNum

    def get_MonthFull(self):
        return self.MonthFull

    def get_MonthAbbr(self):
        return self.MonthAbbr

    def get_QuarterNum(self):
        return self.QuarterNum

    def get_QuarterFull(self):
        return self.QuarterFull

    def get_QuarterAbbr(self):
        return self.QuarterAbbr

    def get_YearAndQuarterNum(self):
        return self.YearAndQuarterNum

    def get_QuarterAbbrAndYear(self):
        return self.QuarterAbbrAndYear

    def get_MonthAbbrAndYear(self):
        return self.MonthAbbrAndYear

    def get_MonthAndYear(self):
        return self.MonthAndYear

    def get_MonthName(self):
        return self.MonthName

    def get_MonthNameAbbr(self):
        return self.MonthNameAbbr

    def get_QuarterAndYear(self):
        return self.QuarterAndYear

    def get_QuarterAndYearAbbr2(self):
        return self.QuarterAndYearAbbr2

    def get_YearAndMonthNum(self):
        return self.YearAndMonthNum

    def set_DateKey(self, DateKey):
        self.DateKey = DateKey

    def set_Year(self, Year):
        self.Year = Year

    def set_MonthNum(self, MonthNum):
        self.MonthNum = MonthNum

    def set_MonthFull(self, MonthFull):
        self.MonthFull = MonthFull

    def set_MonthAbbr(self, MonthAbbr):
        self.MonthAbbr = MonthAbbr

    def set_QuarterNum(self, QuarterNum):
        self.QuarterNum = QuarterNum

    def set_QuarterFull(self, QuarterFull):
        self.QuarterFull = QuarterFull

    def set_QuarterAbbr(self, QuarterAbbr):
        self.QuarterAbbr = QuarterAbbr

    def set_YearAndQuarterNum(self, YearAndQuarterNum):
        self.YearAndQuarterNum = YearAndQuarterNum

    def set_QuarterAbbrAndYear(self, QuarterAbbrAndYear):
        self.QuarterAbbrAndYear = QuarterAbbrAndYear

    def set_MonthAbbrAndYear(self, MonthAbbrAndYear):
        self.MonthAbbrAndYear = MonthAbbrAndYear

    def set_MonthAndYear(self, MonthAndYear):
        self.MonthAndYear = MonthAndYear

    def set_MonthName(self, MonthName):
        self.MonthName = MonthName

    def set_MonthNameAbbr(self, MonthNameAbbr):
        self.MonthNameAbbr = MonthNameAbbr

    def set_QuarterAndYear(self, QuarterAndYear):
        self.QuarterAndYear = QuarterAndYear

    def set_QuarterAndYearAbbr2(self, QuarterAndYearAbbr2):
        self.QuarterAndYearAbbr2 = QuarterAndYearAbbr2

    def set_YearAndMonthNum(self, YearAndMonthNum):
        self.YearAndMonthNum = YearAndMonthNum

    def __str__(self):
        return "DateKey: " + str(self.DateKey) + " Year: " + str(self.Year) + " MonthNum: " + str(self.MonthNum) + " MonthFull: " + str(self.MonthFull) + " MonthAbbr: " + str(self.MonthAbbr) + " QuarterNum: " + str(self.QuarterNum) + " QuarterFull: " + str(self.QuarterFull) + " QuarterAbbr: " + str(self.QuarterAbbr) + " YearAndQuarterNum: " + str(self.YearAndQuarterNum) + " QuarterAbbrAndYear: " + str(self.QuarterAbbrAndYear) + " MonthAbbrAndYear: " + str(self.MonthAbbrAndYear) + " MonthAndYear: " + str(self.MonthAndYear) + " MonthName: " + str(self.MonthName) + " MonthNameAbbr: " + str(self.MonthNameAbbr) + " QuarterAndYear: " + str(self.QuarterAndYear) + " QuarterAndYearAbbr2: " + str(self.QuarterAndYearAbbr2) + " YearAndMonthNum: " + str(self.YearAndMonthNum)

    




# create class invoicelines with attributes InvoiceLineID,InvoiceID,StockID,SalePrice,LineItem
class InvoiceLines:
    def __init__(self, InvoiceLineID, InvoiceID, StockID, SalePrice, LineItem):
        self.InvoiceLineID = InvoiceLineID
        self.InvoiceID = InvoiceID
        self.StockID = StockID
        self.SalePrice = SalePrice
        self.LineItem = LineItem

    def __str__(self):
        return "InvoiceLineID: {}, InvoiceID: {}, StockID: {}, SalePrice: {}, LineItem: {}".format(self.InvoiceLineID, self.InvoiceID, self.StockID, self.SalePrice, self.LineItem)

    def get_InvoiceLineID(self):
        return self.InvoiceLineID

    def get_InvoiceID(self):
        return self.InvoiceID

    def get_StockID(self):
        return self.StockID

    def get_SalePrice(self):
        return self.SalePrice

    def get_LineItem(self):
        return self.LineItem

    def set_InvoiceLineID(self, InvoiceLineID):
        self.InvoiceLineID = InvoiceLineID

    def set_InvoiceID(self, InvoiceID):
        self.InvoiceID = InvoiceID

    def set_StockID(self, StockID):
        self.StockID = StockID

    def set_SalePrice(self, SalePrice):
        self.SalePrice = SalePrice

    def set_LineItem(self, LineItem):
        self.LineItem = LineItem




# create class invoices with attributes     InvoiceID ,InvoiceNumber ,ClientID ,InvoiceDate ,TotalDiscount ,DeliveryCharge ,InvoiceDateKey INT
class Invoices:
    def __init__(self, InvoiceID, InvoiceNumber, ClientID, InvoiceDate, TotalDiscount, DeliveryCharge, InvoiceDateKey):
        self.InvoiceID = InvoiceID
        self.InvoiceNumber = InvoiceNumber
        self.ClientID = ClientID
        self.InvoiceDate = InvoiceDate 
        self.TotalDiscount = TotalDiscount
        self.DeliveryCharge = DeliveryCharge
        self.InvoiceDateKey = InvoiceDateKey

    def get_InvoiceID(self):
        return self.InvoiceID

    def get_InvoiceNumber(self):
        return self.InvoiceNumber

    def get_ClientID(self):
        return self.ClientID

    def get_InvoiceDate(self):
        return self.InvoiceDate

    def get_TotalDiscount(self):
        return self.TotalDiscount

    def get_DeliveryCharge(self):
        return self.DeliveryCharge

    def get_InvoiceDateKey(self):
        return self.InvoiceDateKey

    def set_InvoiceID(self, InvoiceID):
        self.InvoiceID = InvoiceID

    def set_InvoiceNumber(self, InvoiceNumber):
        self.InvoiceNumber = InvoiceNumber

    def set_ClientID(self, ClientID):
        self.ClientID = ClientID

    def set_InvoiceDate(self, InvoiceDate):
        self.InvoiceDate = InvoiceDate

    def set_TotalDiscount(self, TotalDiscount):
        self.TotalDiscount = TotalDiscount

    def set_DeliveryCharge(self, DeliveryCharge):
        self.DeliveryCharge = DeliveryCharge

    def set_InvoiceDateKey(self, InvoiceDateKey):
        self.InvoiceDateKey = InvoiceDateKey

    def __eq__(self, other):
        return self.InvoiceID == other.InvoiceID

    def __str__(self):
        return "InvoiceID: " + str(self.InvoiceID) + " InvoiceNumber: " + str(self.InvoiceNumber) + " ClientID: " + str(
            self.ClientID) + " InvoiceDate: " + str(self.InvoiceDate) + " TotalDiscount: " + str(
            self.TotalDiscount) + " DeliveryCharge: " + str(self.DeliveryCharge) + " InvoiceDateKey: " + str(
            self.InvoiceDateKey)


# create class stock with attributes StockID, Make, Model,ColorID,VehicleType ,CostPrice,SpareParts,LaborCost ,Registration_Date ,Mileage ,PurchaseDate,VehicleAgeInYears

class Stock:
    def __init__(self, StockID, Make, Model,ColorID,VehicleType ,CostPrice,SpareParts,LaborCost ,Registration_Date ,Mileage ,PurchaseDate,VehicleAgeInYears):
        self.StockID = StockID
        self.Make = Make
        self.Model = Model
        self.ColorID = ColorID
        self.VehicleType = VehicleType
        self.CostPrice = CostPrice
        self.SpareParts = SpareParts
        self.LaborCost = LaborCost
        self.Registration_Date = Registration_Date
        self.Mileage = Mileage
        self.PurchaseDate = PurchaseDate
        self.VehicleAgeInYears = VehicleAgeInYears

    def __eq__ (self, other):
        return self.StockID == other.StockID

    def __str__(self):
        return "StockID: " + str(self.StockID) + "\nMake: " + str(self.Make) + "\nModel: " + str(self.Model) + "\nColorID: " + str(self.ColorID) + "\nVehicleType: " + str(self.VehicleType) + "\nCostPrice: " + str(self.CostPrice) + "\nSpareParts: " + str(self.SpareParts) + "\nLaborCost: " + str(self.LaborCost) + "\nRegistration_Date: " + str(self.Registration_Date) + "\nMileage: " + str(self.Mileage) + "\nPurchaseDate: " + str(self.PurchaseDate) + "\nVehicleAgeInYears: " + str(self.VehicleAgeInYears)

    def get_StockID(self):
        return self.StockID
    
    def get_Make(self):
        return self.Make

    def get_Model(self):
        return self.Model

    def get_ColorID(self):
        return self.ColorID

    def get_VehicleType(self):
        return self.VehicleType

    def get_CostPrice(self):
        return self.CostPrice

    def get_SpareParts(self):
        return self.SpareParts

    def get_LaborCost(self):
        return self.LaborCost

    def get_Registration_Date(self):
        return self.Registration_Date

    def get_Mileage(self):
        return self.Mileage

    def get_PurchaseDate(self):
        return self.PurchaseDate

    def get_VehicleAgeInYears(self):
        return self.VehicleAgeInYears

    def set_StockID(self, StockID):
        self.StockID = StockID

    def set_Make(self, Make):
        self.Make = Make

    def set_Model(self, Model):
        self.Model = Model

    def set_ColorID(self, ColorID):
        self.ColorID = ColorID

    def set_VehicleType(self, VehicleType):
        self.VehicleType = VehicleType

    def set_CostPrice(self, CostPrice):
        self.CostPrice = CostPrice

    def set_SpareParts(self, SpareParts):
        self.SpareParts = SpareParts

    def set_LaborCost(self, LaborCost):
        self.LaborCost = LaborCost

    def set_Registration_Date(self, Registration_Date):
        self.Registration_Date = Registration_Date

    def set_Mileage(self, Mileage):
        self.Mileage = Mileage

    def set_PurchaseDate(self, PurchaseDate):
        self.PurchaseDate = PurchaseDate

    def set_VehicleAgeInYears(self, VehicleAgeInYears):
        self.VehicleAgeInYears = VehicleAgeInYears



# create class countries with attributes countryID, countryName, countryISOcode
class Countries:
    def __init__(self, countryID, countryName, countryISOcode):
        self.countryID = countryID
        self.countryName = countryName
        self.countryISOcode = countryISOcode

    
    def get_countryID(self):
        return self.countryID

    def get_countryName(self):
        return self.countryName

    def get_countryISOcode(self):
        return self.countryISOcode

    def set_countryID(self, countryID):
        self.countryID = countryID

    def set_countryName(self, countryName):
        self.countryName = countryName

    def set_countryISOcode(self, countryISOcode):
        self.countryISOcode = countryISOcode

    def __str__ (self):
        return str(self.countryID) + " " + str(self.countryName) + " " + str(self.countryISOcode)

    def __repr__ (self):
        return str(self.countryID) + " " + str(self.countryName) + " " + str(self.countryISOcode)




# create class colors with attributes colorID, color 
class Colors:
    def __init__(self, colorID, color):
        self.colorID = colorID
        self.color = color
    
    def __str__(self):
        return "colorID: " + str(self.colorID) + " color: " + str(self.color)
    
    def __repr__(self):
        return "colorID: " + str(self.colorID) + " color: " + str(self.color)

#define get and set methods for the colors class
    def get_colorID(self):
        return self.colorID
    
    def set_colorID(self, colorID):
        self.colorID = colorID
    
    def get_color(self):
        return self.color

    def set_color(self, color):
        self.color = color







# create class clients with attributes client_id, client_name, address1, address2, town, county, postcode,region, outerpostcode, countryID, ClientType, ClientSize, ClientSince, IsCreditWorthy,IsDealer
class Client:
    def __init__(self, client_id, client_name, address1, address2, town, county, postcode, region, outerpostcode, countryID, ClientType, ClientSize, ClientSince, IsCreditWorthy, IsDealer):
        
        
        
        self.client_id = int(client_id) 
        self.client_name = client_name
        self.address1 = address1
        self.address2 = address2
        self.town = town
        self.county = county
        self.postcode = postcode
        self.region = region
        self.outerpostcode = outerpostcode
        self.countryID = int(countryID)
        self.ClientType = ClientType
        self.ClientSize = ClientSize
        self.ClientSince = ClientSince
        self.IsCreditWorthy = IsCreditWorthy
        self.IsDealer = IsDealer

    # create set and get methods for each attribute
    def get_client_id(self):
        return self.client_id

    def set_client_id(self, client_id):
        self.client_id = client_id

    def get_client_name(self):
        return self.client_name

    def set_client_name(self, client_name):
        self.client_name = client_name

    def get_address1(self):
        return self.address1

    def set_address1(self, address1):
        self.address1 = address1

    def get_address2(self):
        return self.address2

    def set_address2(self, address2):
        self.address2 = address2

    def get_town(self):
        return self.town

    def set_town(self, town):
        self.town = town

    def get_county(self):
        return self.county

    def set_county(self, county):
        self.county = county

    def get_postcode(self):
        return self.postcode

    def set_postcode(self, postcode):
        self.postcode = postcode

    def get_region(self):
        return self.region

    def set_region(self, region):
        self.region = region

    def get_outerpostcode(self):
        return self.outerpostcode

    def set_outerpostcode(self, outerpostcode):
        self.outerpostcode = outerpostcode

    def get_countryID(self):
        return self.countryID

    def set_countryID(self, countryID):
        self.countryID = countryID

    def get_ClientType(self):
        return self.ClientType

    def set_ClientType(self, ClientType):
        self.ClientType = ClientType

    def get_ClientSize(self):
        return self.ClientSize

    def set_ClientSize(self, ClientSize):
        self.ClientSize = ClientSize

    def get_ClientSince(self):
        return self.ClientSince

    def set_ClientSince(self, ClientSince):
        self.ClientSince = ClientSince

    def get_IsCreditWorthy(self):
        return self.IsCreditWorthy

    def set_IsCreditWorthy(self, IsCreditWorthy):
        self.IsCreditWorthy = IsCreditWorthy

    def get_IsDealer(self):
        return self.IsDealer

    def set_IsDealer(self, IsDealer):
        self.IsDealer = IsDealer

    def __str__(self):
        return f"{self.client_id}, {self.client_name}, {self.address1}, {self.address2}, {self.town}, {self.county}, {self.postcode}, {self.region}, {self.outerpostcode}, {self.countryID}, {self.ClientType}, {self.ClientSize}, {self.ClientSince}, {self.IsCreditWorthy}, {self.IsDealer}"

    
def read_clients_xls(file_name):
    # open the file
    workbook = xlrd.open_workbook(file_name)
    # get the first sheet
    worksheet = workbook.sheet_by_index(0)
    # create a list of clients
    clients = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a client object
        client = Client(worksheet.cell_value(row, 0), worksheet.cell_value(row, 1), worksheet.cell_value(row, 2), worksheet.cell_value(row, 3), worksheet.cell_value(row, 4), worksheet.cell_value(row, 5), worksheet.cell_value(row, 6), worksheet.cell_value(row, 7), worksheet.cell_value(row, 8), worksheet.cell_value(row, 9), worksheet.cell_value(row, 10), worksheet.cell_value(row, 11), worksheet.cell_value(row, 12), worksheet.cell_value(row, 13), worksheet.cell_value(row, 14))
        # append the client object to the list
        clients.append(client)
    # return the list of clients
    return clients

def read_colors_xls(file_name):
    # open the file
    workbook = xlrd.open_workbook(file_name)
    # get the first sheet
    worksheet = workbook.sheet_by_index(1)
    # create a list of colors
    colors = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a color object
        color = Colors(worksheet.cell_value(row, 0), worksheet.cell_value(row, 1))
        # append the color object to the list
        colors.append(color)
    # return the list of colors
    return colors

def read_countries_xls(file_name):
    # open the file
    workbook = xlrd.open_workbook(file_name)
    # get the first sheet
    worksheet = workbook.sheet_by_index(2)
    # create a list of countries
    countries = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a country object
        country = Countries(worksheet.cell_value(row, 0), worksheet.cell_value(row, 1),worksheet.cell_value(row, 2) )
        # append the country object to the list
        countries.append(country)
    # return the list of countries
    return countries





def read_stock_xls(filename):
    # open the file
    workbook = xlrd.open_workbook(filename)
    # get the first sheet
    worksheet = workbook.sheet_by_index(3)
    # create a list of stock
    stock = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a stock object with 12 values 
        stock_item = Stock(worksheet.cell_value(row, 0), worksheet.cell_value(row, 1), worksheet.cell_value(row, 2), worksheet.cell_value(row, 3), worksheet.cell_value(row, 4), worksheet.cell_value(row, 5), worksheet.cell_value(row, 6), worksheet.cell_value(row, 7), worksheet.cell_value(row, 8), worksheet.cell_value(row, 9), worksheet.cell_value(row, 10), worksheet.cell_value(row, 11))
        # append the stock object to the list
        stock.append(stock_item)
    # return the list of stock
    return stock

def read_invoices_xls(filename):
    # open the file
    workbook = xlrd.open_workbook(filename)
    # get the first sheet
    worksheet = workbook.sheet_by_index(4)
    # create a list of invoices
    invoices = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a invoice object with 7 values
        invoiceID = int(worksheet.cell_value(row, 0))
        invoice_number = worksheet.cell_value(row, 1)
        invoice_client = int(worksheet.cell_value(row, 2))
        #invoice_date = datetime(worksheet.cell_value(row, 3))

        invoice = Invoices(int(worksheet.cell_value(row, 0)), worksheet.cell_value(row, 1), int(worksheet.cell_value(row, 2)), worksheet.cell_value(row, 3), worksheet.cell_value(row, 4), worksheet.cell_value(row, 5), int(worksheet.cell_value(row, 6))) 
       # append the invoice object to the list
        invoices.append(invoice)
    # return the list of invoices
    return invoices

def read_invoice_lines_xls(filename):
    # open the file
    workbook = xlrd.open_workbook(filename)
    # get the first sheet
    worksheet = workbook.sheet_by_index(5)
    # create a list of invoice lines
    invoice_lines = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a invoice line object with 5 values
        invoice_line = InvoiceLines(worksheet.cell_value(row, 0), worksheet.cell_value(row, 1), worksheet.cell_value(row, 2), worksheet.cell_value(row, 3), worksheet.cell_value(row, 4))
        # append the invoice line object to the list
        invoice_lines.append(invoice_line)
    # return the list of invoice lines
    return invoice_lines


def read_DateDimension_xls(filename):
    # open the file
    workbook = xlrd.open_workbook(filename)
    # get the first sheet
    worksheet = workbook.sheet_by_index(6)
    # create a list of DateDimension
    DateDimensionlst = []
    # loop through the rows
    
    for row in range(1, worksheet.nrows):
        # create a DateDimension object with 17 values
        DateDimension_item = DateDimension(worksheet.cell_value(row, 0), worksheet.cell_value(row, 1), worksheet.cell_value(row, 2), worksheet.cell_value(row, 3), worksheet.cell_value(row, 4), worksheet.cell_value(row, 5), worksheet.cell_value(row, 6), worksheet.cell_value(row, 7), worksheet.cell_value(row, 8), worksheet.cell_value(row, 9), worksheet.cell_value(row, 10), worksheet.cell_value(row, 11), worksheet.cell_value(row, 12), worksheet.cell_value(row, 13), worksheet.cell_value(row, 14), worksheet.cell_value(row, 15), worksheet.cell_value(row, 16))
        # append the DateDimension object to the list
        DateDimensionlst.append(DateDimension_item)
    # return the list of DateDimension
    return DateDimensionlst


def create_stock_list():
   

    # ask user for StockID, Make, Model,ColorID,VehicleType ,CostPrice,SpareParts,LaborCost ,Registration_Date ,Mileage ,PurchaseDate,VehicleAgeInYears
    StockID = input("Enter StockID: ")
    Make = input("Enter Make: ")
    Model = input("Enter Model: ")
    ColorID = input("Enter ColorID: ")
    VehicleType = input("Enter VehicleType: ")
    CostPrice = input("Enter CostPrice: ")
    SpareParts = input("Enter SpareParts: ")
    LaborCost = input("Enter LaborCost: ")
    Registration_Date = input("Enter Registration_Date: ")
    Mileage = input("Enter Mileage: ")
    PurchaseDate = input("Enter PurchaseDate: ")
    VehicleAgeInYears = input("Enter VehicleAgeInYears: ")

    # create a stock object
    stock = Stock(StockID, Make, Model, ColorID, VehicleType, CostPrice, SpareParts, LaborCost, Registration_Date, Mileage, PurchaseDate, VehicleAgeInYears)
    diction = vars(stock)
    lista = diction.items()
    # append the dictionary to dataframe
    df = pd.DataFrame(diction, index=[0])
    print(df)

    book = load_workbook('cargill.xlsx')
    page = book['Stock']
    for row in lista:
        page.append(row)
    book.save('cargill.xlsx')

    
# Create a query that returns the top 3 car brands most sold during first and third quarter of year 2015.

def Convert(a):
    it = iter(a)
    res_dct = dict(zip(it, it))
    return res_dct

def top_3_car_brands2015():
    
    stocklst = read_stock_xls('cargill.xlsx')
    dates = read_DateDimension_xls('cargill.xlsx')
    invoices = read_invoices_xls('cargill.xlsx')
    invoice_lines = read_invoice_lines_xls('cargill.xlsx')

    invoicesids = []
    clientids = []
    invoicedatekeys = []
    makers = []

    invoicelines_stockid = [] 
    for i in invoices:
        
        

        fecha = int(i.InvoiceDateKey)
        fecha = str(fecha)
        year = fecha[0:4]
        month = fecha[4:6]
        day = fecha[6:8]
        str_date = year + '-' + month + '-' + day
        # str_date to datetime
        date = datetime.strptime(str_date, '%Y-%m-%d')
        
        quarter = pd.Timestamp(date).quarter
        if quarter == 1 and int(year) == 2015 or quarter == 3 and int(year) == 2015:
            invoicedatekeys.append(str_date)
            invoicesids.append(i.InvoiceID)
            clientids.append(i.ClientID)

    for item in invoice_lines:
        if item.InvoiceID in invoicesids:
            invoicelines_stockid.append(int(item.StockID))

    # get top 3 most frequent values in invoicelines_stockid
    

    #print(invoicelines_stockid)
    for item in stocklst:
        if item.StockID in invoicelines_stockid:
            makers.append(item.Make)
            
    diccionario = {'stockid':invoicelines_stockid,'maker':makers,'invoicedatekey':invoicedatekeys}

    df = pd.DataFrame(diccionario)
    #print(df)
    n = int(3)
    freq = df['maker'].value_counts()[:n].index.tolist()
    print("most frequent 2015",freq)


def top_3_car_brands(years):
    
    stocklst = read_stock_xls('cargill.xlsx')
    dates = read_DateDimension_xls('cargill.xlsx')
    invoices = read_invoices_xls('cargill.xlsx')
    invoice_lines = read_invoice_lines_xls('cargill.xlsx')
    colorslst = read_colors_xls('cargill.xlsx')

    invoicesidsq1 = []
    clientidsq1 = []
    invoicedatekeysq1 = []
    invoicesidsq2 = []
    clientidsq2 = []
    invoicedatekeysq2 = []
    invoicesidsq3 = []
    clientidsq3 = []
    invoicedatekeysq3 = []
    invoicesidsq4 = []
    clientidsq4 = []
    invoicedatekeysq4 = []

    
    years = ['2012','2013','2014','2015']

    clrss_q1 = []
    clrss_q2 = []
    clrss_q3 = []
    clrss_q4 = []

    clrsname_q1 = []
    clrsname_q2 = []
    clrsname_q3 = []
    clrsname_q4 = []

    invoicelines_stockidq1 = []
    invoicelines_stockidq2 = []
    invoicelines_stockidq3 = []
    invoicelines_stockidq4 = []
    for iyear in years: 
        for i in invoices:
            
            

            fecha = int(i.InvoiceDateKey)
            fecha = str(fecha)
            year = fecha[0:4]
            month = fecha[4:6]
            day = fecha[6:8]
            str_date = year + '-' + month + '-' + day
            # str_date to datetime
            date = datetime.strptime(str_date, '%Y-%m-%d')
            
            quarter = pd.Timestamp(date).quarter
            if quarter == 1 and int(year) == int(iyear) :
                invoicedatekeysq1.append(str_date)
                invoicesidsq1.append(i.InvoiceID)
                clientidsq1.append(i.ClientID)
            elif quarter == 2 and int(year) == int(iyear):
                invoicedatekeysq2.append(str_date)
                invoicesidsq2.append(i.InvoiceID)
                clientidsq2.append(i.ClientID)
            elif quarter == 3 and int(year) == int(iyear):
                invoicedatekeysq3.append(str_date)
                invoicesidsq3.append(i.InvoiceID)
                clientidsq3.append(i.ClientID)
            elif quarter == 4 and int(year) == int(iyear):
                invoicedatekeysq4.append(str_date)
                invoicesidsq4.append(i.InvoiceID)
                clientidsq4.append(i.ClientID)

        for item in invoice_lines:
            if item.InvoiceID in invoicesidsq1:
                invoicelines_stockidq1.append(int(item.StockID))
            elif item.InvoiceID in invoicesidsq2:
                invoicelines_stockidq2.append(int(item.StockID))
            elif item.InvoiceID in invoicesidsq3:
                invoicelines_stockidq3.append(int(item.StockID))
            elif item.InvoiceID in invoicesidsq4:
                invoicelines_stockidq4.append(int(item.StockID))
        

        # get top 3 most frequent values in invoicelines_stockid
        

        #print(invoicelines_stockid)
        for item in stocklst:
            if item.StockID in invoicelines_stockidq1:
                clrss_q1.append(item.ColorID)
            if item.StockID in invoicelines_stockidq2:
                clrss_q2.append(item.ColorID)
            if item.StockID in invoicelines_stockidq3:
                clrss_q3.append(item.ColorID)
            if item.StockID in invoicelines_stockidq4:
                clrss_q4.append(item.ColorID)

        for item in colorslst:
            if item.colorID in clrss_q1:
                clrsname_q1.append(item.color)
            if item.colorID in clrss_q2:
                clrsname_q2.append(item.color)
            if item.colorID in clrss_q3:
                clrsname_q3.append(item.color)
            if item.colorID in clrss_q4:
                clrsname_q4.append(item.color)
                
        diccionario_q1 = {'color_q1':clrsname_q1}
        df_q1 = pd.DataFrame(diccionario_q1)
        diccionario_q2 = {'color_q2':clrsname_q2}
        df_q2 = pd.DataFrame(diccionario_q2)
        diccionario_q3 = {'color_q3':clrsname_q3}
        df_q3 = pd.DataFrame(diccionario_q3)
        diccionario_q4 = {'color_q4':clrsname_q4}
        df_q4 = pd.DataFrame(diccionario_q4)
        #print(df)
        n = int(3)
        freq = df_q1['color_q1'].value_counts()[:n].index.tolist()
        print("most frequent color "+iyear+" Q1",freq)
        freq = df_q2['color_q2'].value_counts()[:n].index.tolist()
        print("most frequent color "+iyear+" Q2",freq)
        freq = df_q3['color_q3'].value_counts()[:n].index.tolist()
        print("most frequent color "+iyear+" Q3",freq)
        freq = df_q4['color_q4'].value_counts()[:n].index.tolist()
        print("most frequent color "+iyear+" Q4",freq)

       




if __name__ == "__main__":
    top_3_car_brands(['2012','2013','2014','2015'])
    top_3_car_brands2015()
    #create_stock_list()
    StockID = input("Enter StockID: ")
    Make = input("Enter Make: ")
    Model = input("Enter Model: ")
    ColorID = input("Enter ColorID: ")
    VehicleType = input("Enter VehicleType: ")
    CostPrice = input("Enter CostPrice: ")
    SpareParts = input("Enter SpareParts: ")
    LaborCost = input("Enter LaborCost: ")
    Registration_Date = input("Enter Registration_Date: ")
    Mileage = input("Enter Mileage: ")
    PurchaseDate = input("Enter PurchaseDate: ")
    VehicleAgeInYears = input("Enter VehicleAgeInYears: ")
    barcos = Stock_Barco(StockID, Make, Model, ColorID, VehicleType, CostPrice, SpareParts, LaborCost, Registration_Date, Mileage, PurchaseDate, VehicleAgeInYears)
    barcos.create_stock_list()
    
